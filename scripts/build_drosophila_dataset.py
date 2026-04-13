#!/usr/bin/env python3

from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency 'openpyxl'. Run `python3 -m pip install -r requirements.txt` "
        "from the project root, then retry the build."
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw" / "drosophila"
PROCESSED_DIR = ROOT / "data" / "processed"
META_DIR = ROOT / "data" / "meta"

WORKBOOK_PATH = RAW_DIR / "source_data" / "Source_Data" / "Source_Data.xlsx"
ZIP_PATH = RAW_DIR / "41467_2022_30975_MOESM16_ESM.zip"

COMBINED_CSV_PATH = PROCESSED_DIR / "drosophila_lifespan_all.csv"
MANIFEST_PATH = META_DIR / "drosophila_dataset_manifest.json"

DEFAULT_SITE = "ALL"
CONTROL_GROUP = "AL"
TREATMENT_GROUP = "DR"
SPECIES = "D. melanogaster"

PUBLICATION = {
    "pmid": "35672419",
    "doi": "10.1038/s41467-022-30975-4",
    "title": "Dietary restriction and the transcription factor clock delay eye aging to extend lifespan in Drosophila melanogaster.",
    "author": "Hodge et al.",
    "year": 2022,
    "journal": "Nature Communications",
}

COHORT_SPECS = [
    {
        "sheet_name": "5a. W1118_LD_DD_Lifespan",
        "cohort": "w1118_dd",
        "label": "w1118 in constant darkness",
        "short_label": "w1118 DD",
        "secondary_label": "AL vs DR",
        "strain": "w1118",
        "condition": "constant darkness",
        "blocks": [
            {"group_header_row": 2, "replicate_row": 3, "data_end_row": 44},
            {"group_header_row": 45, "replicate_row": 46, "data_end_row": 81},
            {"group_header_row": 82, "replicate_row": 83, "data_end_row": 120},
        ],
        "source_data_label": "Figure 5a (DD)",
    },
    {
        "sheet_name": "5a. W1118_LD_DD_Lifespan",
        "cohort": "w1118_ld",
        "label": "w1118 in 12:12 light-dark",
        "short_label": "w1118 LD",
        "secondary_label": "AL vs DR",
        "strain": "w1118",
        "condition": "12:12 light-dark",
        "blocks": [
            {"group_header_row": 121, "replicate_row": 122, "data_end_row": 161},
            {"group_header_row": 162, "replicate_row": 163, "data_end_row": 203},
            {"group_header_row": 204, "replicate_row": 205, "data_end_row": 240},
        ],
        "source_data_label": "Figure 5a (LD)",
    },
    {
        "sheet_name": "Sup7a. CantonS_LD_DD_Lifespan",
        "cohort": "canton_s_ld",
        "label": "Canton-S in 12:12 light-dark",
        "short_label": "Canton-S LD",
        "secondary_label": "AL vs DR",
        "strain": "Canton-S",
        "condition": "12:12 light-dark",
        "blocks": [
            {"group_header_row": 2, "replicate_row": 3, "data_end_row": 48},
            {"group_header_row": 98, "replicate_row": 99, "data_end_row": 145},
            {"group_header_row": 98, "replicate_row": 196, "data_end_row": 236},
        ],
        "source_data_label": "Supplementary Figure 7a (LD)",
    },
    {
        "sheet_name": "Sup7a. CantonS_LD_DD_Lifespan",
        "cohort": "canton_s_dd",
        "label": "Canton-S in constant darkness",
        "short_label": "Canton-S DD",
        "secondary_label": "AL vs DR",
        "strain": "Canton-S",
        "condition": "constant darkness",
        "blocks": [
            {"group_header_row": 50, "replicate_row": 51, "data_end_row": 96},
            {"group_header_row": 147, "replicate_row": 148, "data_end_row": 194},
            {"group_header_row": 238, "replicate_row": 239, "data_end_row": 279},
        ],
        "source_data_label": "Supplementary Figure 7a (DD)",
    },
    {
        "sheet_name": "5b. ninaE_null_Lifespan",
        "cohort": "ninae17_ld",
        "label": "ninaE17 mutant in 12:12 light-dark",
        "short_label": "ninaE17",
        "secondary_label": "AL vs DR",
        "strain": "ninaE17",
        "condition": "12:12 light-dark",
        "blocks": [
            {"group_header_row": 1, "replicate_row": 2, "data_end_row": None},
        ],
        "source_data_label": "Figure 5b",
    },
    {
        "sheet_name": "5c. Rh3_null_Lifespan",
        "cohort": "rh3_null_ld",
        "label": "rh3 mutant in 12:12 light-dark",
        "short_label": "rh3 null",
        "secondary_label": "AL vs DR",
        "strain": "rh3 null",
        "condition": "12:12 light-dark",
        "blocks": [
            {"group_header_row": 1, "replicate_row": 2, "data_end_row": None},
        ],
        "source_data_label": "Figure 5c",
    },
    {
        "sheet_name": "5d. Rh4_null_Lifespan",
        "cohort": "rh4_null_ld",
        "label": "rh4 mutant in 12:12 light-dark",
        "short_label": "rh4 null",
        "secondary_label": "AL vs DR",
        "strain": "rh4 null",
        "condition": "12:12 light-dark",
        "blocks": [
            {"group_header_row": 1, "replicate_row": 2, "data_end_row": None},
        ],
        "source_data_label": "Figure 5d",
    },
    {
        "sheet_name": "5e. Rh6_null_Lifespan",
        "cohort": "rh6_null_ld",
        "label": "rh6 mutant in 12:12 light-dark",
        "short_label": "rh6 null",
        "secondary_label": "AL vs DR",
        "strain": "rh6 null",
        "condition": "12:12 light-dark",
        "blocks": [
            {"group_header_row": 1, "replicate_row": 2, "data_end_row": None},
        ],
        "source_data_label": "Figure 5e",
    },
    {
        "sheet_name": "5f. Gqalpha_null_Lifespan",
        "cohort": "gqalpha_null_ld",
        "label": "Gqalpha mutant in 12:12 light-dark",
        "short_label": "Gqalpha null",
        "secondary_label": "AL vs DR",
        "strain": "Gqalpha null",
        "condition": "12:12 light-dark",
        "blocks": [
            {"group_header_row": 2, "replicate_row": 3, "data_end_row": None},
        ],
        "source_data_label": "Figure 5f",
    },
    {
        "sheet_name": "Sup7b. TRP365_mutant_Lifespan",
        "cohort": "trp365_dd",
        "label": "TRP365 mutant in constant darkness",
        "short_label": "TRP365 DD",
        "secondary_label": "AL vs DR",
        "strain": "TRP365 mutant",
        "condition": "constant darkness",
        "blocks": [
            {"group_header_row": 41, "replicate_row": 42, "data_end_row": 79},
            {"group_header_row": 120, "replicate_row": 121, "data_end_row": 154},
        ],
        "source_data_label": "Supplementary Figure 7b (DD)",
    },
]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_group_label(value: object) -> str | None:
    text = str(value or "").strip().upper()
    if not text:
        return None
    if "DR" in text:
        return TREATMENT_GROUP
    if "AL" in text:
        return CONTROL_GROUP
    return None


def parse_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        parsed = int(round(float(value)))
    except (TypeError, ValueError):
        return None
    return parsed


def get_column_specs(
    sheet_rows: list[tuple[object, ...]],
    *,
    group_header_row: int,
    replicate_row: int,
) -> list[dict[str, object]]:
    current_group: str | None = None
    column_specs: list[dict[str, object]] = []
    max_column = max((len(row) for row in sheet_rows), default=0)
    group_cells = sheet_rows[group_header_row - 1] if len(sheet_rows) >= group_header_row else ()
    replicate_cells = sheet_rows[replicate_row - 1] if len(sheet_rows) >= replicate_row else ()

    for col_index in range(2, max_column + 1):
        group_value = group_cells[col_index - 1] if len(group_cells) >= col_index else None
        replicate_value = (
            replicate_cells[col_index - 1] if len(replicate_cells) >= col_index else None
        )
        current_group = normalize_group_label(group_value) or current_group
        replicate = parse_int(replicate_value)

        if current_group is None or replicate is None:
            continue

        column_specs.append(
            {
                "col_index": col_index,
                "group": current_group,
                "replicate": replicate,
            }
        )

    if not column_specs:
        raise SystemExit(
            "Unable to determine AL/DR replicate columns from parsed worksheet rows"
        )

    return column_specs


def build_group_meta(
    *,
    cohort: str,
    strain: str,
    condition: str,
) -> dict[str, dict[str, object]]:
    condition_sentence = f" under {condition} conditions" if condition else ""

    return {
        f"{cohort}::{CONTROL_GROUP}": {
            "cohort": cohort,
            "group": CONTROL_GROUP,
            "control_group": CONTROL_GROUP,
            "label": "Ad libitum control",
            "description": (
                "A cohort-matched high-yeast ad libitum control condition for fruit flies. "
                f"In this experiment, {strain} flies were reared on 5% yeast extract{condition_sentence}."
            ),
            "is_control": True,
            "species": SPECIES,
            "strain": strain,
            "compound": "ad_libitum",
            "compound_display_name": "Ad libitum control",
            "control_compound": "ad_libitum",
            "concentration": "5",
            "units": "% yeast extract",
            "condition": condition,
            "publication": PUBLICATION,
        },
        f"{cohort}::{TREATMENT_GROUP}": {
            "cohort": cohort,
            "group": TREATMENT_GROUP,
            "control_group": CONTROL_GROUP,
            "label": "Dietary restriction",
            "description": (
                "A reduced-yeast dietary restriction condition tested against the matched ad libitum cohort. "
                f"In this experiment, {strain} flies were reared on 0.5% yeast extract{condition_sentence}."
            ),
            "is_control": False,
            "species": SPECIES,
            "strain": strain,
            "compound": "dietary_restriction",
            "compound_display_name": "Dietary restriction",
            "control_compound": "ad_libitum",
            "concentration": "0.5",
            "units": "% yeast extract",
            "condition": condition,
            "publication": PUBLICATION,
        },
    }


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Missing source workbook: {WORKBOOK_PATH}")

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    META_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)

    fieldnames = [
        "cohort",
        "site",
        "sex",
        "id",
        "group",
        "control_group",
        "age_days",
        "dead",
        "status",
        "species",
        "strain",
        "compound",
        "compound_display_name",
        "control_compound",
        "concentration",
        "units",
        "condition",
        "is_control",
        "age_initiation_months",
    ]

    combined_rows: list[dict[str, object]] = []
    cohort_meta_by_name: dict[str, dict[str, object]] = {}
    group_meta_by_key: dict[str, dict[str, object]] = {}
    raw_sources: list[dict[str, object]] = []

    for spec in COHORT_SPECS:
        worksheet = workbook[spec["sheet_name"]]
        sheet_rows = list(worksheet.iter_rows(values_only=True))

        group_counts: Counter[str] = Counter()
        replicate_counts: Counter[str] = Counter()
        observed_ages: list[int] = []
        event_index = 0

        for block_index, block in enumerate(spec["blocks"], start=1):
            group_header_row = int(block["group_header_row"])
            replicate_row = int(block["replicate_row"])
            data_end_row = int(block["data_end_row"] or len(sheet_rows))
            column_specs = get_column_specs(
                sheet_rows,
                group_header_row=group_header_row,
                replicate_row=replicate_row,
            )

            for row_index, row_values in enumerate(
                sheet_rows[replicate_row:data_end_row],
                start=replicate_row + 1,
            ):
                age_days = parse_int(row_values[0] if row_values else None)
                if age_days is None:
                    continue

                for column_spec in column_specs:
                    col_index = int(column_spec["col_index"])
                    cell_value = row_values[col_index - 1] if len(row_values) >= col_index else None
                    cell_count = parse_int(cell_value)
                    if cell_count is None or cell_count <= 0:
                        continue

                    group = str(column_spec["group"])
                    replicate = int(column_spec["replicate"])
                    compound = (
                        "ad_libitum" if group == CONTROL_GROUP else "dietary_restriction"
                    )
                    compound_display_name = (
                        "Ad libitum control"
                        if group == CONTROL_GROUP
                        else "Dietary restriction"
                    )

                    for death_copy in range(cell_count):
                        event_index += 1
                        combined_rows.append(
                            {
                                "cohort": spec["cohort"],
                                "site": DEFAULT_SITE,
                                "sex": "all",
                                "id": (
                                    f"FLY-{spec['cohort']}-B{block_index}-{group}-"
                                    f"R{replicate}-{row_index}-{death_copy + 1}"
                                ),
                                "group": group,
                                "control_group": CONTROL_GROUP,
                                "age_days": age_days,
                                "dead": "true",
                                "status": "dead",
                                "species": SPECIES,
                                "strain": spec["strain"],
                                "compound": compound,
                                "compound_display_name": compound_display_name,
                                "control_compound": "ad_libitum",
                                "concentration": "5" if group == CONTROL_GROUP else "0.5",
                                "units": "% yeast extract",
                                "condition": spec["condition"],
                                "is_control": str(group == CONTROL_GROUP).lower(),
                                "age_initiation_months": "",
                            }
                        )

                    group_counts[group] += cell_count
                    replicate_counts[f"B{block_index}:{group}:R{replicate}"] += cell_count
                    observed_ages.extend([age_days] * cell_count)

        if not group_counts[CONTROL_GROUP] or not group_counts[TREATMENT_GROUP]:
            raise SystemExit(
                f"Missing AL/DR event rows after parsing worksheet {spec['sheet_name']!r}"
            )

        cohort_meta_by_name[spec["cohort"]] = {
            "label": spec["label"],
            "short_label": spec["short_label"],
            "secondary_label": spec["secondary_label"],
            "sheet_name": spec["sheet_name"],
            "source_data_label": spec["source_data_label"],
            "strain": spec["strain"],
            "condition": spec["condition"],
            "publication": PUBLICATION,
            "min_age_days": min(observed_ages),
            "max_age_days": max(observed_ages),
        }
        group_meta_by_key.update(
            build_group_meta(
                cohort=spec["cohort"],
                strain=spec["strain"],
                condition=spec["condition"],
            )
        )
        raw_sources.append(
            {
                "cohort": spec["cohort"],
                "sheet_name": spec["sheet_name"],
                "source_data_label": spec["source_data_label"],
                "row_count": len(sheet_rows),
                "column_count": max((len(row) for row in sheet_rows), default=0),
                "block_count": len(spec["blocks"]),
                "normalized_row_count": group_counts[CONTROL_GROUP] + group_counts[TREATMENT_GROUP],
                "group_counts": dict(group_counts),
                "replicate_counts": dict(replicate_counts),
                "min_age_days": min(observed_ages),
                "max_age_days": max(observed_ages),
            }
        )

    combined_rows.sort(
        key=lambda row: (
            str(row["cohort"]),
            str(row["group"]),
            int(row["age_days"]),
            str(row["id"]),
        )
    )

    with COMBINED_CSV_PATH.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(combined_rows)

    cohort_order = [spec["cohort"] for spec in COHORT_SPECS]
    manifest = {
        "generated_at_utc": utc_now_iso(),
        "latest_public_cohort_downloaded": cohort_order[-1],
        "latest_public_release_label": "Nat Commun 2022",
        "combined_csv": str(COMBINED_CSV_PATH.relative_to(ROOT)),
        "cohort_order": cohort_order,
        "cohort_meta_by_name": cohort_meta_by_name,
        "site_meta": {DEFAULT_SITE: "Nature source data"},
        "group_meta_by_key": group_meta_by_key,
        "counts": {
            "rows": len(combined_rows),
            "cohorts": len(cohort_order),
            "intervention_groups": sum(
                1 for meta in group_meta_by_key.values() if not meta["is_control"]
            ),
            "dead_events": len(combined_rows),
            "censored_rows": 0,
            "sex_counts": {"all": len(combined_rows)},
        },
        "raw_sources": [
            {
                "source_zip": str(ZIP_PATH.relative_to(ROOT)) if ZIP_PATH.exists() else None,
                "source_workbook": str(WORKBOOK_PATH.relative_to(ROOT)),
                "sheets": raw_sources,
            }
        ],
        "notes": [
            "The curated workbook subset covers AL-versus-DR lifespan experiments from wild-type and phototransduction-mutant Drosophila cohorts.",
            "The source workbook stores per-day death counts by replicate/vial; these counts were expanded into one row per fly to match the app's survival schema.",
            "The paper describes these lifespan assays as female-fly populations; the normalized dataset uses sex=all so the app renders a single combined panel.",
        ],
    }

    with MANIFEST_PATH.open("w") as handle:
        json.dump(manifest, handle, indent=2)
        handle.write("\n")

    print(
        f"Wrote {COMBINED_CSV_PATH.relative_to(ROOT)} with {len(combined_rows)} rows and "
        f"{MANIFEST_PATH.relative_to(ROOT)}"
    )


if __name__ == "__main__":
    main()
