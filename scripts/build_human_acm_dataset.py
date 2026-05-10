#!/usr/bin/env python3

from __future__ import annotations

import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw" / "human_acm"
META_DIR = ROOT / "data" / "meta"
WORKBOOK_PATH = RAW_DIR / "ACEL-23-e14334-s001.xlsx"
DOWNLOAD_MANIFEST_PATH = RAW_DIR / "human_acm_download_manifest.json"
OUTPUT_PATH = META_DIR / "human_acm_dataset_manifest.json"

COHORT = "HUMAN_UKB_ACM_MORIN_2024"

PUBLICATION = {
    "title": "Association between prescription drugs and all-cause mortality risk in the UK population",
    "author": "Morin et al.",
    "year": 2024,
    "journal": "Aging Cell",
    "pmid": "39364726",
    "doi": "10.1111/acel.14334",
    "pmc": "PMC11634711",
    "url": "https://pmc.ncbi.nlm.nih.gov/articles/PMC11634711/",
}

SUPPLEMENT_WORKBOOK_URL = (
    "https://pmc.ncbi.nlm.nih.gov/articles/instance/11634711/bin/"
    "ACEL-23-e14334-s001.xlsx"
)
SUPPLEMENT_ARCHIVE_URL = (
    "https://pmc.ncbi.nlm.nih.gov/articles/instance/11634711/bin/"
    "ACEL-23-e14334-s002.zip"
)
SUPPLEMENT_DATASET_LINKS = {
    "Data Table 2": {
        "label": "Dataset 2",
        "table": "Data Table 2",
        "url": SUPPLEMENT_WORKBOOK_URL,
        "detail": "No-concentration drug ACM result table in ACEL-23-e14334-s001.xlsx.",
    },
    "Data Table 3": {
        "label": "Dataset 3",
        "table": "Data Table 3",
        "url": SUPPLEMENT_WORKBOOK_URL,
        "detail": "Dose/formulation ACM result table in ACEL-23-e14334-s001.xlsx.",
    },
    "Data Table 4": {
        "label": "Dataset 4",
        "table": "Data Table 4",
        "url": SUPPLEMENT_WORKBOOK_URL,
        "detail": "Medication-class ACM result table in ACEL-23-e14334-s001.xlsx.",
    },
    "Data Table 5": {
        "label": "Dataset 5",
        "table": "Data Table 5",
        "url": SUPPLEMENT_WORKBOOK_URL,
        "detail": "Harmonized drug-name table in ACEL-23-e14334-s001.xlsx.",
    },
}

UKB_PROJECT_PUBLICATION = {
    "title": "Systematic identification of potential lifespan-modulating drugs and long-term health outcomes in the UK Biobank",
    "author": "Ristow lab",
    "year": 2026,
    "journal": "UK Biobank approved research project 1246676",
    "url": "https://www.ukbiobank.ac.uk/projects/systematic-identification-of-potential-lifespan-modulating-drugs-and-long-term-health-outcomes-in-the-uk-biobank/",
}

SOURCE_CATALOG = [
    {
        "key": "aging-cell-2024-uk-biobank-acm-supplement",
        "label": "UK Biobank 406-medication mortality screen",
        "summary": (
            "Public Aging Cell 2024 supplement from a medication-wide observational "
            "analysis of UK Biobank prescription-drug users and all-cause mortality."
        ),
        "finding": (
            "The source pool contains all 406 public Data Table 2 no-concentration "
            "medication rows with N>=500 plus the six Figure 5 / Data Table 4 "
            "medication-class rows; the site view displays rows with reported P values "
            "and ranks them as 1 minus the reported ACM hazard ratio."
        ),
        "dataset_links": [
            SUPPLEMENT_DATASET_LINKS["Data Table 2"],
            SUPPLEMENT_DATASET_LINKS["Data Table 4"],
            SUPPLEMENT_DATASET_LINKS["Data Table 5"],
            {
                "label": "Data S2 archive",
                "url": SUPPLEMENT_ARCHIVE_URL,
                "detail": "Supplementary archive distributed with the paper.",
            },
        ],
        "publication": PUBLICATION,
    },
    {
        "key": "uk-biobank-project-1246676",
        "label": "UK Biobank lifespan-modulating drug project",
        "summary": (
            "Official UK Biobank project page for approved research project 1246676, "
            "focused on medication exposures, all-cause mortality, multimorbidity, "
            "and candidate geroprotective drugs."
        ),
        "finding": (
            "Included as project provenance for the human Biobank view; the ACM rows "
            "come from the public Aging Cell supplement, not participant-level UKB data."
        ),
        "publication": UKB_PROJECT_PUBLICATION,
    },
]

ROW_OVERRIDES: dict[str, dict[str, str]] = {
    "ATORVASTATIN": {
        "display": "Atorvastatin",
        "description": "A statin drug that lowers LDL cholesterol by inhibiting HMG-CoA reductase.",
        "condition": "UK Biobank general-population prescription users",
    },
    "ATORVASTATIN 20MG CAPS OR TABS": {
        "display": "Atorvastatin 20 mg",
        "description": "A statin drug that lowers LDL cholesterol by inhibiting HMG-CoA reductase.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "NAPROXEN": {
        "display": "Naproxen",
        "description": "A nonsteroidal anti-inflammatory drug (NSAID) used for pain, fever, and inflammation.",
        "condition": "UK Biobank general-population prescription users",
    },
    "NAPROXEN 500MG CAPS OR TABS": {
        "display": "Naproxen 500 mg",
        "description": "A nonsteroidal anti-inflammatory drug (NSAID) used for pain, fever, and inflammation.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "SILDENAFIL": {
        "display": "Sildenafil",
        "description": "A PDE5 inhibitor used for erectile dysfunction and pulmonary arterial hypertension.",
        "condition": "UK Biobank general-population prescription users",
    },
    "SILDENAFIL 50MG CAPS OR TABS": {
        "display": "Sildenafil 50 mg",
        "description": "A PDE5 inhibitor used for erectile dysfunction and pulmonary arterial hypertension.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "OTOMIZE": {
        "display": "Otomize",
        "description": "An ear spray combining neomycin, dexamethasone, and acetic acid for inflammatory outer-ear infections.",
        "condition": "UK Biobank general-population prescription users",
    },
    "OTOMIZE EAR SPRAY": {
        "display": "Otomize ear spray",
        "description": "An ear spray combining neomycin, dexamethasone, and acetic acid for inflammatory outer-ear infections.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "ESTRADIOL": {
        "display": "Estradiol",
        "description": "The primary endogenous estrogen, used clinically in hormone replacement and local estrogen therapies.",
        "condition": "UK Biobank female-only prescription users",
    },
    "VAGIFEM": {
        "display": "Vagifem",
        "description": "An estradiol vaginal tablet used as local estrogen therapy for menopausal urogenital symptoms.",
        "condition": "UK Biobank female-only prescription users",
    },
    "VAGIFEM 10MG VAGINAL TABS": {
        "display": "Vagifem 10 mg",
        "description": "An estradiol vaginal tablet used as local estrogen therapy for menopausal urogenital symptoms.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "ESTRIOL": {
        "display": "Estriol",
        "description": "A naturally occurring estrogen used in some local hormone-therapy preparations.",
        "condition": "UK Biobank female-only prescription users",
    },
    "OVESTIN": {
        "display": "Ovestin",
        "description": "An estriol vaginal cream used as local estrogen therapy for menopausal urogenital symptoms.",
        "condition": "UK Biobank female-only prescription users",
    },
    "AVAXIM VACCINE SUSPENSION FOR INJECTION": {
        "display": "Avaxim",
        "description": "An inactivated hepatitis A vaccine used to prevent hepatitis A infection.",
        "condition": "UK Biobank vaccine prescription users",
    },
    "AVAXIM VACCINE SUSPENSION FOR INJECTION 0.5ML PRE-FILLED SYRINGES": {
        "display": "Avaxim 0.5 ml pre-filled syringe",
        "description": "An inactivated hepatitis A vaccine used to prevent hepatitis A infection.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "LYMECYCLINE": {
        "display": "Lymecycline",
        "description": "A tetracycline-class antibiotic used for acne and other susceptible bacterial infections.",
        "condition": "UK Biobank general-population prescription users",
    },
    "LYMECYCLINE 408MG CAPS OR TABS": {
        "display": "Lymecycline 408 mg",
        "description": "A tetracycline-class antibiotic used for acne and other susceptible bacterial infections.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "ESTRADERM": {
        "display": "Estraderm",
        "description": "An estradiol transdermal patch used as estrogen replacement therapy.",
        "condition": "UK Biobank female-only prescription users",
    },
    "REVAXIS VACCINE SUSPENSION FOR INJECTION": {
        "display": "Revaxis",
        "description": "A booster vaccine against diphtheria, tetanus, and poliomyelitis.",
        "condition": "UK Biobank vaccine prescription users",
    },
    "REVAXIS VACCINE SUSPENSION FOR INJECTION 0.5ML PRE-FILLED SYRINGES": {
        "display": "Revaxis 0.5 ml pre-filled syringe",
        "description": "A booster vaccine against diphtheria, tetanus, and poliomyelitis.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "MARVELON": {
        "display": "Marvelon",
        "description": "A combined oral contraceptive containing desogestrel and ethinylestradiol.",
        "condition": "UK Biobank female-only prescription users",
    },
    "MARVELON CAPS OR TABS": {
        "display": "Marvelon caps or tabs",
        "description": "A combined oral contraceptive containing desogestrel and ethinylestradiol.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "TERBINAFINE 250MG CAPS OR TABS": {
        "display": "Terbinafine 250 mg",
        "description": "An allylamine antifungal drug used for fungal skin and nail infections.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "PREMARIN 0.625MG CAPS OR TABS": {
        "display": "Premarin 0.625 mg",
        "description": "Also called conjugated estrogens, an estrogen-therapy drug used for menopausal symptoms and other hormone indications.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "MELOXICAM 7.5MG CAPS OR TABS": {
        "display": "Meloxicam 7.5 mg",
        "description": "A nonsteroidal anti-inflammatory drug (NSAID) used for pain and inflammation, especially arthritis.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "TADALAFIL 10MG CAPS OR TABS": {
        "display": "Tadalafil 10 mg",
        "description": "A PDE5 inhibitor used for erectile dysfunction, pulmonary arterial hypertension, and urinary symptoms from benign prostatic hyperplasia.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "ACE": {
        "display": "ACE inhibitors",
        "description": "A renin-angiotensin-system medication class used for hypertension, heart failure, kidney protection, and related cardiovascular indications.",
        "condition": "UK Biobank ACE-inhibitor class prescription users",
    },
    "FLECAINIDE 100MG CAPS OR TABS": {
        "display": "Flecainide 100 mg",
        "description": "A class IC antiarrhythmic drug and sodium-channel blocker used for selected heart rhythm disorders.",
        "condition": "UK Biobank dose-specific prescription users",
    },
    "Metformin": {
        "display": "Metformin",
        "description": "A glucose-lowering biguanide medicine used for type 2 diabetes and cardiometabolic-risk management.",
        "condition": "UK Biobank metformin prescription users",
    },
    "Estrogens": {
        "display": "Estrogen therapy",
        "description": "A hormone-therapy class using estrogenic compounds for menopausal and other estrogen-deficiency indications.",
        "condition": "UK Biobank estrogen-class prescription users",
    },
    "PDE5": {
        "display": "PDE5 inhibitors",
        "description": "A drug class that inhibits phosphodiesterase type 5 and is used for erectile dysfunction and pulmonary hypertension.",
        "condition": "UK Biobank medication-class prescription users",
    },
    "SGLT2": {
        "display": "SGLT2 inhibitors",
        "description": "A class of diabetes and cardiorenal drugs that lower blood glucose by increasing urinary glucose excretion.",
        "condition": "UK Biobank medication-class prescription users",
    },
    "Statins": {
        "display": "Statins",
        "description": "A cholesterol-lowering drug class that inhibits HMG-CoA reductase.",
        "condition": "UK Biobank medication-class prescription users",
    },
}

ROW_NOTES: dict[str, str] = {
    "Marvelon": "Marvelon was named among the increased-lifespan drugs and ranks first because Data Table 2 reports the lowest ACM hazard ratio among the decreased-mortality rows.",
    "Revaxis": "Revaxis was named among the decreased-mortality drugs and ranks above SGLT2i because its individual-drug HR implies a larger ACM reduction.",
    "Avaxim": "Avaxim is included because Data Table 2 reports a lower ACM hazard ratio for this prescription-drug row.",
    "Flecainide 100 mg": "Flecainide 100 mg is included from Data Table 3 because the concentration-specific row passed FDR, while the parent-drug aggregate was not a no-concentration hit.",
    "Tadalafil 10 mg": "Tadalafil 10 mg is included from Data Table 3 because the concentration-specific row passed FDR, while the parent-drug aggregate was not a no-concentration hit.",
    "Meloxicam 7.5 mg": "Meloxicam 7.5 mg is included from Data Table 3 because the concentration-specific row passed FDR, while the parent-drug aggregate was not a no-concentration hit.",
    "Premarin 0.625 mg": "Premarin 0.625 mg is included from Data Table 3 because the concentration-specific row passed FDR, while the parent-drug aggregate was not a no-concentration hit.",
    "Terbinafine 250 mg": "Terbinafine 250 mg is included from Data Table 3 because the concentration-specific row passed FDR, while the parent-drug aggregate was not a no-concentration hit.",
    "Estrogen therapy": "Estrogen therapy is included as a class-level Biobank signal separate from the individual estradiol row, and ranks by its reported ACM hazard-ratio effect.",
    "Estraderm": "Estraderm ranks above the broader estrogen class because its individual-drug HR implies a larger ACM reduction.",
    "Estradiol": "Estradiol is retained as a separate individual-drug row and ranks by its reported ACM hazard-ratio effect.",
    "Sildenafil": "Sildenafil ranks by its UK Biobank ACM hazard-ratio effect, with no non-Biobank evidence streams contributing to the order.",
    "Lymecycline": "Lymecycline is included because Figure 2 reports a lower ACM hazard ratio for this prescription-drug row.",
    "Vagifem": "Vagifem is retained as a separate individual-drug row because the article reports its ACM HR independently from the broader estrogen class.",
    "Estriol": "Estriol is retained as a separate individual-drug row because the article reports its ACM HR independently from the broader estrogen class.",
    "Otomize": "Otomize is included because Figure 2 and Data Table 2 report a lower ACM hazard ratio for this prescription-drug row.",
    "PDE5 inhibitors": "PDE5 inhibitors are included as a class-level Biobank signal separate from the individual sildenafil row.",
    "Naproxen": "Naproxen ranks by its reported ACM hazard-ratio effect, with no non-Biobank evidence streams contributing to the order.",
    "Atorvastatin": "Atorvastatin remains in the human view only through its UK Biobank ACM effect; trial meta-analysis and Mendelian evidence are intentionally excluded.",
    "SGLT2 inhibitors": "SGLT2 inhibitors remain the strongest medication-class signal, although individual-drug rows such as Marvelon and Revaxis have larger reported ACM effects.",
    "Statins": "Statins are included as a class-level Biobank row because the article reports HR 0.97, but this should be read as borderline because the CI reaches 1.00.",
}

DRUG_DETAIL_PATTERNS: list[tuple[str, dict[str, str]]] = [
    (
        r"\b(CIALIS|TADALAFIL)\b",
        {
            "description": "A PDE5 inhibitor used for erectile dysfunction, pulmonary arterial hypertension, and urinary symptoms from benign prostatic hyperplasia.",
            "condition": "UK Biobank PDE5 inhibitor prescription users",
        },
    ),
    (
        r"\b(CERAZETTE|MICRONOR|NORIDAY|NORETHISTERONE|OVRANETTE|CILEST|MICROGYNON|LOGYNON)\b",
        {
            "description": "A hormonal contraceptive or progestogen medicine used for contraception, menstrual symptoms, or other hormone-related indications.",
            "condition": "UK Biobank female-only prescription users",
        },
    ),
    (
        r"\b(FEMOSTON|CONJUGATED OESTROGENS|ELLESTE|CLIMESSE|ESTRACOMBI|PREMARIN|KLIOVANCE|NUVELLE|TIBOLONE|PREMIQUE|EVOREL|ORTHO-GYNEST|KLIOFEM|CLIMAGEST|PREMPAK|CLIMAVAL|LIVIAL)\b",
        {
            "description": "A hormone-therapy preparation used for menopausal symptoms, local estrogen replacement, or other estrogen/progestogen indications.",
            "condition": "UK Biobank female-only prescription users",
        },
    ),
    (
        r"\b(REPLENSMD)\b",
        {
            "description": "A vaginal moisturiser used for vaginal dryness and menopausal urogenital symptoms.",
            "condition": "UK Biobank female-only prescription users",
        },
    ),
    (
        r"\b(CELLUVISC|CARBOMER|CARMELLOSE|LACRI-LUBE|HYPROMELLOSE|OPTIVE|VISCOTEARS)\b",
        {
            "description": "An ocular lubricant or artificial-tear preparation used for dry-eye symptoms and eye-surface protection.",
            "condition": "UK Biobank ocular lubricant users",
        },
    ),
    (
        r"\b(BIMATOPROST|TRAVOPROST|TIMOLOL|BRINZOLAMIDE|XALATAN|LATANOPROST)\b",
        {
            "description": "A glaucoma or ocular-hypertension eye medicine used to lower intraocular pressure.",
            "condition": "UK Biobank glaucoma eye-drop users",
        },
    ),
    (
        r"\b(NEDOCROMIL|AVAMYS|PSEUDOEPHEDRINE|SODIUM CROMOGLICATE|OPTICROM|FLIXONASE|NASONEX|BECONASE|MOMETASONE|MONTELUKAST)\b",
        {
            "description": "An allergy, rhinitis, asthma, or airway medicine used for allergic inflammation, nasal symptoms, or respiratory control.",
            "condition": "UK Biobank allergy or airway prescription users",
        },
    ),
    (
        r"\b(OMEPRAZOLE|LANSOPRAZOLE|PANTOPRAZOLE|ESOMEPRAZOLE|RABEPRAZOLE|RANITIDINE|CIMETIDINE|GAVISCON|PEPTAC|SODIUM ALGINATE|POTASSIUM BICARBONATE)\b",
        {
            "description": "An acid-suppression, H2-blocker, or alginate antacid medicine used for reflux, ulcer disease, dyspepsia, or gastroprotection.",
            "condition": "UK Biobank acid-suppression or antacid prescription users",
        },
    ),
    (
        r"\b(RAMIPRIL|LISINOPRIL|ENALAPRIL|PERINDOPRIL|CANDESARTAN|LOSARTAN|IRBESARTAN|VALSARTAN|OLMESARTAN|ACE)\b",
        {
            "description": "A renin-angiotensin-system medicine used for hypertension, heart failure, kidney protection, or other cardiovascular indications.",
            "condition": "UK Biobank RAAS prescription users",
        },
    ),
    (
        r"\b(CLONIDINE|CO-TENIDONE)\b",
        {
            "description": "An antihypertensive medicine or cardiovascular combination used for blood-pressure control and related cardiovascular indications.",
            "condition": "UK Biobank antihypertensive prescription users",
        },
    ),
    (
        r"\b(GLICLAZIDE|GLIMEPIRIDE|GLIPIZIDE|METFORMIN|SITAGLIPTIN|LINAGLIPTIN|PIOGLITAZONE|ROSIGLITAZONE)\b",
        {
            "description": "A glucose-lowering diabetes medicine used for type 2 diabetes and related cardiometabolic-risk management.",
            "condition": "UK Biobank non-insulin diabetes prescription users",
        },
    ),
    (
        r"\b(SIMVADOR|EZETIMIBE|FENOFIBRATE|BEZAFIBRATE)\b",
        {
            "description": "A lipid-lowering medicine used to manage cholesterol, triglycerides, or broader cardiovascular risk.",
            "condition": "UK Biobank lipid-lowering prescription users",
        },
    ),
    (
        r"\b(ATORVASTATIN|SIMVASTATIN|ROSUVASTATIN|PRAVASTATIN|FLUVASTATIN)\b",
        {
            "description": "A statin drug that lowers LDL cholesterol by inhibiting HMG-CoA reductase.",
            "condition": "UK Biobank statin prescription users",
        },
    ),
    (
        r"\b(ORLISTAT)\b",
        {
            "description": "A gastrointestinal lipase inhibitor used for weight management in obesity or overweight with risk factors.",
            "condition": "UK Biobank weight-management prescription users",
        },
    ),
    (
        r"\b(ASPIRIN|CLOPIDOGREL|DIPYRIDAMOLE|TICAGRELOR|PRASUGREL)\b",
        {
            "description": "An antiplatelet or vascular-risk medicine used to reduce thrombotic events in cardiovascular or cerebrovascular disease contexts.",
            "condition": "UK Biobank antiplatelet prescription users",
        },
    ),
    (
        r"\b(CITALOPRAM|SERTRALINE|FLUOXETINE|PAROXETINE|MIRTAZAPINE|AMITRIPTYLINE|NORTRIPTYLINE|DULOXETINE|VENLAFAXINE|TRAZODONE|LOFEPRAMINE|DOSULEPIN|ESCITALOPRAM)\b",
        {
            "description": "An antidepressant or neuromodulatory medicine used for depression, anxiety, neuropathic pain, sleep, or related indications.",
            "condition": "UK Biobank antidepressant or neuromodulator prescription users",
        },
    ),
    (
        r"\b(SOLIFENACIN|OXYBUTYNIN|TOLTERODINE|DOXAZOSIN|TAMSULOSIN|ALFUZOSIN|FINASTERIDE|DUTASTERIDE|MIRABEGRON)\b",
        {
            "description": "A urologic medicine used for overactive bladder, urinary symptoms, or prostate-related urinary-flow indications.",
            "condition": "UK Biobank urologic prescription users",
        },
    ),
    (
        r"\b(THYROXINE|CARBIMAZOLE)\b",
        {
            "description": "A thyroid medicine used for thyroid hormone replacement or treatment of overactive thyroid disease.",
            "condition": "UK Biobank thyroid prescription users",
        },
    ),
    (
        r"\b(ALLOPURINOL|COLCHICINE|FEBUXOSTAT)\b",
        {
            "description": "A urate-lowering or gout medicine used for gout, hyperuricemia, or inflammatory crystal-disease prevention.",
            "condition": "UK Biobank gout or urate-lowering prescription users",
        },
    ),
    (
        r"\b(FEXOFENADINE|CETIRIZINE|LORATADINE|DESLORATADINE|LEVOCETIRIZINE|TERFENADINE|CHLORPHENAMINE|PROMETHAZINE|HYDROXYZINE)\b",
        {
            "description": "An antihistamine medicine used for allergy, urticaria, itching, or sedating symptom-control indications.",
            "condition": "UK Biobank antihistamine prescription users",
        },
    ),
    (
        r"\b(EPIPEN)\b",
        {
            "description": "An adrenaline auto-injector used for emergency treatment of severe allergic reactions and anaphylaxis.",
            "condition": "UK Biobank adrenaline auto-injector users",
        },
    ),
    (
        r"\b(QUININE)\b",
        {
            "description": "A quinoline medicine historically used for malaria and sometimes prescribed for nocturnal leg cramps.",
            "condition": "UK Biobank quinine prescription users",
        },
    ),
    (
        r"\b(MALARONE)\b",
        {
            "description": "An atovaquone-proguanil antimalarial medicine used for malaria prevention or treatment.",
            "condition": "UK Biobank antimalarial prescription users",
        },
    ),
    (
        r"\b(AVIVA TESTING STRIPS|CONTOUR|ACTIVE TESTING STRIPS|ONETOUCH|ACCU-CHEK|FREESTYLE|MICROLET LANCETS|MULTICLIX LANCETS|SOFTCLIX|SHARPSAFE|SHARPSGUARD|GLUCORX|MEDISENSE|DIASTIX|COMPACT TESTING STRIPS|ADVANTAGE II BIOSENSOR)\b",
        {
            "description": "A glucose-monitoring supply row used by people checking blood glucose, commonly in diabetes care.",
            "condition": "UK Biobank diabetes-monitoring supply users",
        },
    ),
    (
        r"\b(SUMATRIPTAN|RIZATRIPTAN|ZOLMITRIPTAN|NARATRIPTAN|PIZOTIFEN)\b",
        {
            "description": "A triptan migraine medicine that activates serotonin 5-HT1 receptors to treat acute migraine attacks.",
            "condition": "UK Biobank migraine prescription users",
        },
    ),
    (
        r"\b(ENZIRA|INFLUVAC|FLUARIX|PNEUMOVAX|PREVENAR|ZOSTAVAX|TYPHIM|HAVRIX|IMUVAC|INFLUENZA VACCINE|BEGRIVAC)\b",
        {
            "description": "A vaccine prescription row marking immunization against influenza, pneumococcal disease, shingles, or other infections.",
            "condition": "UK Biobank vaccine prescription users",
        },
    ),
    (
        r"\b(SCHERIPROCT|PROCTOSEDYL|XYLOPROCT|ANUSOL HC)\b",
        {
            "description": "An anorectal symptom medicine used for haemorrhoids, anal irritation, or local rectal inflammation.",
            "condition": "UK Biobank anorectal prescription users",
        },
    ),
    (
        r"\b(OILATUM|CETRABEN|EPADERM|DOUBLEBASE|E45|BETNOVATE|HYDROCORTISONE|CLOTRIMAZOLE|POLYTAR|AQUEOUS|DIPROBASE|FUSIDIC ACID|BETAMETHASONE VALERATE|TERBINAFINE .+CREAM|BETAMETHASONE|AVEENO|TRIMOVATE|TIMODINE|CAPASAL|DERMOL|CLOBETASONE|CLOBETASOL|DERMOVATE|DAKTACORT|CANESTEN|EUMOVATE|HYDROMOL|DOVOBET|ZEROBASE|CALCIPOTRIOL|FUCIBET|FUCIDIN|GENTISONE|LOCORTEN VIOFORM|NASEPTIN|MOMETASONE|CLIMB|SCHERIPROCT|PROCTOSEDYL|XYLOPROCT|ANUSOL HC)\b",
        {
            "description": "A dermatology medicine, emollient, topical steroid, or antifungal used for eczema, dermatitis, fungal infection, or skin-barrier care.",
            "condition": "UK Biobank dermatology prescription users",
        },
    ),
    (
        r"\b(VARENICLINE)\b",
        {
            "description": "A nicotinic-receptor partial agonist used for smoking cessation.",
            "condition": "UK Biobank smoking-cessation prescription users",
        },
    ),
    (
        r"\b(DONEPEZIL|RIVASTIGMINE|GALANTAMINE|MEMANTINE)\b",
        {
            "description": "A dementia-treatment medicine used for cognitive symptoms in Alzheimer's disease and related neurodegenerative disorders.",
            "condition": "UK Biobank dementia-treatment prescription users",
        },
    ),
    (
        r"\b(ROPINIROLE|RASAGILINE|AMANTADINE|LEVODOPA|MADOPAR|SINEMET|CO-CARELDOPA|CO-BENELDOPA|PRAMIPEXOLE)\b",
        {
            "description": "A Parkinson's disease or movement-disorder medicine acting on dopaminergic motor pathways.",
            "condition": "UK Biobank Parkinson or movement-disorder prescription users",
        },
    ),
    (
        r"\b(MORPHINE SULFATE|ORAMORPH|MST CONTINUS|ZOMORPH|FENTANYL|OXYCODONE|OXYNORM|OXYCONTIN|BUPRENORPHINE|TRAMADOL|DIHYDROCODEINE|CODEINE PHOSPHATE|CO-CODAMOL|CO-DYDRAMOL|CO-PROXAMOL|CODEINE|TYLEX|ZAPAIN)\b",
        {
            "description": "An opioid analgesic used for moderate-to-severe pain, often in severe illness, cancer pain, postoperative, or palliative-care contexts.",
            "condition": "UK Biobank opioid or severe-pain prescription users",
        },
    ),
    (
        r"\b(LORAZEPAM|CLONAZEPAM|DIAZEPAM|TEMAZEPAM|ZOPICLONE|NITRAZEPAM|CHLORDIAZEPOXIDE|ZOLPIDEM)\b",
        {
            "description": "A benzodiazepine or related sedative medicine used for anxiety, insomnia, muscle spasm, or seizure-related indications.",
            "condition": "UK Biobank sedative or anxiolytic prescription users",
        },
    ),
    (
        r"\b(RISPERIDONE|QUETIAPINE|OLANZAPINE|HALOPERIDOL|AMISULPRIDE|ARIPIPRAZOLE)\b",
        {
            "description": "An antipsychotic medicine that modulates dopamine and serotonin signaling for psychosis, bipolar disorder, agitation, or related indications.",
            "condition": "UK Biobank antipsychotic prescription users",
        },
    ),
    (
        r"\b(EPILIM|SODIUM VALPROATE|VALPROIC ACID|LEVETIRACETAM|LAMOTRIGINE|CARBAMAZEPINE|TEGRETOL|PHENYTOIN|PREGABALIN|GABAPENTIN|TOPIRAMATE)\b",
        {
            "description": "An anti-seizure medicine used for epilepsy, seizure prevention, neuropathic pain, or mood-stabilizing indications.",
            "condition": "UK Biobank anti-seizure prescription users",
        },
    ),
    (
        r"\b(AMIODARONE|IVABRADINE|FLECAINIDE|SOTALOL)\b",
        {
            "description": "A cardiac rate or rhythm medicine used for arrhythmia, angina, or heart-failure symptom-control indications.",
            "condition": "UK Biobank cardiac rhythm prescription users",
        },
    ),
    (
        r"\b(MYCOPHENOLATE|AZATHIOPRINE|TACROLIMUS|CICLOSPORIN|CYCLOSPORIN|METHOTREXATE|PREDNISOLONE|HYDROXYCHLOROQUINE|SULFASALAZINE)\b",
        {
            "description": "An immunosuppressive medicine used after organ transplantation or for autoimmune and inflammatory diseases.",
            "condition": "UK Biobank immunosuppressant prescription users",
        },
    ),
    (
        r"\b(EXEMESTANE|LETROZOLE|ANASTROZOLE|TAMOXIFEN)\b",
        {
            "description": "A cancer endocrine-therapy medicine used mainly for hormone-receptor-positive breast cancer.",
            "condition": "UK Biobank cancer endocrine-therapy prescription users",
        },
    ),
    (
        r"\b(DIGOXIN)\b",
        {
            "description": "A cardiac glycoside used for selected heart-failure and atrial-fibrillation rate-control settings.",
            "condition": "UK Biobank cardiac glycoside prescription users",
        },
    ),
    (
        r"\b(SPIRONOLACTONE|EPLERENONE|FUROSEMIDE|FRUSEMIDE|BUMETANIDE|TORASEMIDE|AMILORIDE|CO-AMILOFRUSE|BENDROFLUAZIDE|BENDROFLUMETHIAZIDE|INDAPAMIDE)\b",
        {
            "description": "A diuretic or mineralocorticoid-pathway medicine used for heart failure, edema, hypertension, or cardiorenal fluid overload.",
            "condition": "UK Biobank cardiorenal or diuretic prescription users",
        },
    ),
    (
        r"\b(NIFEDIPINE|ADALAT|AMLODIPINE|DILTIAZEM|VERAPAMIL|LERCANIDIPINE|FELODIPINE)\b",
        {
            "description": "A calcium-channel blocker used for hypertension, angina, or other cardiovascular indications.",
            "condition": "UK Biobank calcium-channel blocker prescription users",
        },
    ),
    (
        r"\b(ISOTARD|ISOSORBIDE|GLYCERYL TRINITRATE|NITROGLYCERIN|NICORANDIL)\b",
        {
            "description": "A nitrate vasodilator used for angina and other ischemic-heart-disease symptom control.",
            "condition": "UK Biobank nitrate vasodilator prescription users",
        },
    ),
    (
        r"\b(CARVEDILOL|BISOPROLOL|ATENOLOL|METOPROLOL|NEBIVOLOL|PROPRANOLOL|HALF INDERAL)\b",
        {
            "description": "A beta-blocker medicine used for heart failure, hypertension, angina, rhythm control, or other cardiovascular indications.",
            "condition": "UK Biobank beta-blocker prescription users",
        },
    ),
    (
        r"\b(FLUDROCORTISONE)\b",
        {
            "description": "A mineralocorticoid corticosteroid used for adrenal insufficiency, orthostatic hypotension, or salt-wasting indications.",
            "condition": "UK Biobank mineralocorticoid prescription users",
        },
    ),
    (
        r"\b(SALBUTAMOL|SALMETEROL|SERETIDE|SEREVENT|VENTOLIN|SALAMOL|TERBUTALINE|FLUTICASONE|IPRATROPIUM|COMBIVENT|THEOPHYLLINE|UNIPHYLLIN|PHYLLOCONTIN|CARBOCISTEINE|BECLOMETASONE|BUDESONIDE|FORMOTEROL|SYMBICORT|TIOTROPIUM|VOLUMATIC|AEROCHAMBER|QVAR|CLENIL|FOSTAIR|BECOTIDE|PEAK FLOW METER)\b",
        {
            "description": "A respiratory medicine used for asthma, COPD, bronchospasm, airway inflammation, or mucus-clearance indications.",
            "condition": "UK Biobank respiratory prescription users",
        },
    ),
    (
        r"\b(CREON|PANCREATIN)\b",
        {
            "description": "A pancreatic enzyme replacement used to support digestion in pancreatic exocrine insufficiency.",
            "condition": "UK Biobank pancreatic enzyme prescription users",
        },
    ),
    (
        r"\b(ALVERINE|MEBEVERINE|PEPPERMINT OIL|COLOFAC|COLPERMIN|BUSCOPAN)\b",
        {
            "description": "A gastrointestinal antispasmodic or peppermint-oil preparation used for irritable-bowel-type abdominal cramping and bowel symptoms.",
            "condition": "UK Biobank gastrointestinal antispasmodic users",
        },
    ),
    (
        r"\b(ASACOL|SULFASALAZINE)\b",
        {
            "description": "An anti-inflammatory medicine used for inflammatory bowel disease or related immune-inflammatory conditions.",
            "condition": "UK Biobank intestinal anti-inflammatory prescription users",
        },
    ),
    (
        r"\b(CYCLIZINE|METOCLOPRAMIDE|ONDANSETRON|PROCHLORPERAZINE|DOMPERIDONE|HYOSCINE BUTYLBROMIDE|LOPERAMIDE|BETAHISTINE|CINNARIZINE)\b",
        {
            "description": "An antiemetic or prokinetic medicine used for nausea, vomiting, vestibular symptoms, or gastrointestinal motility indications.",
            "condition": "UK Biobank antiemetic prescription users",
        },
    ),
    (
        r"\b(GLUCOSAMINE|GAMOLENIC|EFAMAST|OLIVE OIL)\b",
        {
            "description": "A supplement or non-prescription preparation used for joint symptoms, skin symptoms, or general supportive care.",
            "condition": "UK Biobank supplement or supportive-care users",
        },
    ),
    (
        r"\b(THIAMINE|VITAMIN B|VITAMINS|FORCEVAL|FOLIC ACID|CYANOCOBALAMIN|HYDROXOCOBALAMIN|FERROUS|GLUCOSAMINE|GAMOLENIC|EFAMAST|OLIVE OIL)\b",
        {
            "description": "A vitamin replacement or multivitamin preparation used for deficiency, malnutrition risk, alcohol-related care, or supplementation.",
            "condition": "UK Biobank vitamin replacement prescription users",
        },
    ),
    (
        r"\b(ALFACALCIDOL|STRONTIUM RANELATE|COLECALCIFEROL|CALCIUM CARBONATE|ADCAL-D3\w*|CALCEOS|CALCICHEW|ALENDRONIC ACID|RISEDRONATE|FULTIUM|ACCRETE)\b",
        {
            "description": "A bone-mineral or active vitamin D medicine used for calcium, phosphate, osteoporosis, or renal bone-disease indications.",
            "condition": "UK Biobank bone-mineral prescription users",
        },
    ),
    (
        r"\b(DOCUSATE|LAXIDO|LACTULOSE|SENNA|MACROGOL|FYBOGEL|ISPAGHULA|BISACODYL|MOVICOL)\b",
        {
            "description": "A laxative or stool-softening medicine used for constipation and bowel-care indications.",
            "condition": "UK Biobank bowel-care prescription users",
        },
    ),
    (
        r"\b(SODIUM BICARBONATE)\b",
        {
            "description": "An alkalinizing medicine used for metabolic acidosis, renal tubular acidosis, or related acid-base indications.",
            "condition": "UK Biobank acid-base or renal prescription users",
        },
    ),
    (
        r"\b(CHLORHEXIDINE|DIFFLAM|BENZYDAMINE)\b",
        {
            "description": "An oral antiseptic, mouthwash, or anti-inflammatory rinse used for dental, mouth, throat, or mucosal-care indications.",
            "condition": "UK Biobank oral-care prescription users",
        },
    ),
    (
        r"\b(NYSTATIN|AZITHROMYCIN|FLUCONAZOLE|CO-AMOXICLAV|AMOXICILLIN|DOXYCYCLINE|CLARITHROMYCIN|CEFALEXIN|CEFRADINE|CEFACLOR|CIPROFLOXACIN|TRIMETHOPRIM|NITROFURANTOIN|FLUCLOXACILLIN|ERYTHROMYCIN|PHENOXYMETHYLPENICILLIN|PENICILLIN V|METRONIDAZOLE|CO-FLUAMPICIL|OXYTETRACYCLINE|CHLORAMPHENICOL|ACICLOVIR|KETOCONAZOLE|NIZORAL|ITRACONAZOLE|MICONAZOLE|AMOROLFINE|TERBINAFINE|FUCITHALMIC|OTOSPORIN)\b",
        {
            "description": "An antimicrobial or antifungal medicine used for bacterial, fungal, or other susceptible infections.",
            "condition": "UK Biobank antimicrobial prescription users",
        },
    ),
    (
        r"\b(TRANEXAMIC ACID)\b",
        {
            "description": "An antifibrinolytic medicine used to reduce heavy bleeding, including heavy menstrual bleeding or surgical bleeding contexts.",
            "condition": "UK Biobank antifibrinolytic prescription users",
        },
    ),
    (
        r"\b(WARFARIN|RIVAROXABAN|APIXABAN|DABIGATRAN|EDOXABAN)\b",
        {
            "description": "An anticoagulant medicine used to reduce clotting risk in atrial fibrillation, venous thromboembolism, or other thrombotic-risk settings.",
            "condition": "UK Biobank anticoagulant prescription users",
        },
    ),
    (
        r"\b(PARACETAMOL|NEFOPAM|BACLOFEN|MELOXICAM|MEFENAMIC ACID|CELECOXIB|ETORICOXIB|DIETHYLAMINE SALICYLATE|DEPO-MEDRONE|FENBID|PIROXICAM|IBUPROFEN|DICLOFENAC|ROFECOXIB|VOLTAROL|ALGESAL|MOVELAT|KETOPROFEN|INDOMETACIN|ARTHROTEC|CAPSAICIN|LIDOCAINE)\b",
        {
            "description": "A non-opioid analgesic, anti-inflammatory, local anaesthetic, or muscle-spasm medicine used for pain or musculoskeletal symptoms.",
            "condition": "UK Biobank non-opioid pain or spasm prescription users",
        },
    ),
    (
        r"\b(INSULIN|NOVOMIX|NOVORAPID|HUMULIN|LANTUS|LEVEMIR|TOUJEO|APIDRA)\b",
        {
            "description": "An insulin preparation used for diabetes glucose control, often marking insulin-treated or more advanced diabetes.",
            "condition": "UK Biobank insulin-treated diabetes prescription users",
        },
    ),
    (
        r"\b(LITHIUM CARBONATE|LITHIUM CITRATE|PRIADEL)\b",
        {
            "description": "A mood-stabilizing lithium salt used for bipolar disorder and related psychiatric indications.",
            "condition": "UK Biobank lithium-prescribed participants",
        },
    ),
    (
        r"\b(FEMULEN)\b",
        {
            "description": "A progestogen-only contraceptive medicine used for hormonal contraception.",
            "condition": "UK Biobank female-only prescription users",
        },
    ),
]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean_number(value: Any) -> float | int | str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return value
    return str(value).strip()


def to_float(value: Any) -> float | None:
    if value in (None, "", "NA"):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def to_int(value: Any) -> int | None:
    parsed = to_float(value)
    if parsed is None:
        return None
    return int(parsed)


def is_true(value: Any) -> bool:
    if value is True:
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"true", "yes", "1"}
    return False


def round_float(value: float | None, digits: int = 6) -> float | None:
    if value is None:
        return None
    return round(float(value), digits)


def fmt_float(value: float | None, digits: int = 2) -> str:
    if value is None:
        return "NA"
    return f"{value:.{digits}f}"


def format_p(value: float | None) -> str:
    if value is None:
        return "NA"
    if value < 0.001:
        return f"{value:.2e}"
    return f"{value:.4f}"


def slugify(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", value.upper()).strip("_")
    return re.sub(r"_+", "_", slug)


def base_drug_name(raw_drug: str) -> str:
    base = raw_drug.upper()
    base = re.sub(r"\s+\d+(\.\d+)?\s*(MG|ML|MCG|G)\b.*$", "", base)
    base = base.replace(" VACCINE SUSPENSION FOR INJECTION", "")
    base = base.replace(" EAR SPRAY", "")
    base = base.replace(" CAPS OR TABS", "")
    base = base.replace(" VAGINAL TABS", "")
    return base.strip()


def prettify_drug_display(raw_drug: str) -> str:
    display = raw_drug.title()
    replacements = [
        (r"\bMst\b", "MST"),
        (r"\bLa\b", "LA"),
        (r"\bXl\b", "XL"),
        (r"\bCopd\b", "COPD"),
        (r"\bHiv\b", "HIV"),
        (r"(\d)(Mgs)\b", r"\1 mg"),
        (r"(\d)(Mg)\b", r"\1 mg"),
        (r"(\d)(Mcg)\b", r"\1 mcg"),
        (r"(\d)(Ml)\b", r"\1 ml"),
        (r"(\d)(Units)\b", r"\1 units"),
        (r"(\d)(Nanogram)\b", r"\1 nanogram"),
        (r"(\d)(Microgram)\b", r"\1 microgram"),
        (r"\bMgs\b", "mg"),
        (r"\bMg\b", "mg"),
        (r"\bMcg\b", "mcg"),
        (r"\bMl\b", "ml"),
        (r"\bUnits\b", "units"),
        (r"\bCaps Or Tabs\b", "caps or tabs"),
        (r"\bTabs\b", "tabs"),
        (r"\bGastro-Resistant\b", "gastro-resistant"),
        (r"\bPre-Filled\b", "pre-filled"),
        (r"\bOral Solution\b", "oral solution"),
        (r"\bOral Suspension\b", "oral suspension"),
        (r"\bTransdermal Patches\b", "transdermal patches"),
        (r"\bNebuliser\b", "nebuliser"),
        (r"\bMouthwash\b", "mouthwash"),
        (r"\bInhaler\b", "inhaler"),
        (r"\bDose\b", "dose"),
        (r"\bHour\b", "hour"),
        (r"\bActuation\b", "actuation"),
        (r"\bInhalation\b", "inhalation"),
    ]

    for pattern, replacement in replacements:
        display = re.sub(pattern, replacement, display)

    display = re.sub(
        r"\s+(\d+(?:\.\d+)?\s+mg\s+gastro-resistant tabs)\s+\d+(?:\.\d+)?\s+mg\s+caps or tabs$",
        r" \1",
        display,
        flags=re.IGNORECASE,
    )

    return re.sub(r"\s+", " ", display).strip()


def rows_from_sheet(workbook, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    output = []

    for row in rows:
        if not row or row[0] is None:
            continue
        output.append(
            {
                header: clean_number(row[index] if index < len(row) else None)
                for index, header in enumerate(headers)
            },
        )

    return output


def is_fdr_significant_acm_row(row: dict[str, Any]) -> bool:
    hr = to_float(row.get("HR"))
    n = to_int(row.get("N"))
    adjusted_p_n500 = to_float(row.get("Padj N>500"))
    return (
        hr is not None
        and hr != 1
        and n is not None
        and n >= 500
        and adjusted_p_n500 is not None
        and adjusted_p_n500 <= 0.05
    )


def screened_acm_rows(workbook) -> list[dict[str, Any]]:
    selected: list[dict[str, Any]] = []

    for row in rows_from_sheet(workbook, "Data Table 2"):
        hr = to_float(row.get("HR"))
        n = to_int(row.get("N"))
        p_value = to_float(row.get("P"))
        adjusted_p_n500 = to_float(row.get("Padj N>500"))
        if hr is not None and n is not None and n >= 500:
            selected.append(
                {
                    **row,
                    "source_table": "Data Table 2",
                    "source_kind": "drug",
                    "fdr_selection_adjusted_p": adjusted_p_n500,
                    "is_fdr_significant": is_fdr_significant_acm_row(row),
                    "is_statistically_significant": p_value is not None and p_value <= 0.05,
                },
            )

    for row in rows_from_sheet(workbook, "Data Table 4"):
        hr = to_float(row.get("HR"))
        n = to_int(row.get("N"))
        p_value = to_float(row.get("P"))
        if hr is not None and n is not None:
            selected.append(
                {
                    **row,
                    "source_table": "Data Table 4",
                    "source_kind": "class",
                    "fdr_selection_adjusted_p": None,
                    "is_fdr_significant": False,
                    "is_statistically_significant": p_value is not None and p_value <= 0.05,
                },
            )

    return selected


def row_display_info(raw_drug: str) -> dict[str, str]:
    override = ROW_OVERRIDES.get(raw_drug)
    if override:
        return override

    display = prettify_drug_display(raw_drug)
    for pattern, details in DRUG_DETAIL_PATTERNS:
        if re.search(pattern, raw_drug, re.IGNORECASE):
            return {
                "display": display,
                "description": details["description"],
                "condition": details["condition"],
            }

    return {
        "display": display,
        "description": "A UK Biobank prescription-medication exposure from the public mortality-screen supplement.",
        "condition": "UK Biobank prescription users",
    }


def source_label(source_table: str) -> str:
    if source_table == "Data Table 2":
        return "no-concentration drug analysis"
    if source_table == "Data Table 3":
        return "dose/formulation analysis"
    if source_table == "Data Table 4":
        return "medication-class analysis"
    return "supplement analysis"


def biobank_effect_label(display: str, source_kind: str, hr: float | None) -> str:
    direction = "higher" if hr is not None and hr > 1 else "lower"
    if source_kind == "class":
        return f"The UK Biobank Figure 5 ACM analysis reports {display} with {direction} all-cause mortality risk."
    if source_kind == "dose/formulation":
        return f"The UK Biobank concentration analysis highlighted {display} with {direction} all-cause mortality risk."
    return f"The UK Biobank 406-medication screen reports {display} with {direction} all-cause mortality risk."


def overall_effect_label(
    sort_index: int,
    source_kind: str,
    effect_percent: float | None,
    is_fdr_significant: bool,
) -> str:
    if not is_fdr_significant:
        return "Screened ACM row"

    if effect_percent is not None and effect_percent < 0:
        if source_kind == "class":
            return "Higher class-level ACM effect"
        if source_kind == "dose/formulation":
            return "Higher dose-specific ACM effect"
        return "Higher ACM signal"

    if sort_index == 1:
        return "Largest individual ACM effect"
    if source_kind == "class":
        return "Class-level ACM effect"
    if source_kind == "dose/formulation":
        return "Dose-specific ACM effect"
    return "Named ACM effect"


def build_group_meta(row: dict[str, Any], sort_index: int) -> tuple[str, dict[str, Any]]:
    raw_drug = str(row["Drug"])
    row_source_kind = str(row.get("source_kind", "drug"))
    display_info = row_display_info(raw_drug)
    if row_source_kind == "class" and raw_drug == "Metformin":
        display_info = {
            **display_info,
            "display": "Metformin (Figure 5)",
            "condition": "UK Biobank Figure 5 metformin prescription users",
        }
    display = display_info["display"]
    hr = to_float(row.get("HR"))
    lower = to_float(row.get("lower95"))
    upper = to_float(row.get("upper95"))
    p_value = to_float(row.get("P"))
    fdr_value = (
        to_float(row.get("fdr_selection_adjusted_p"))
        or to_float(row.get("Padj N>500"))
        or to_float(row.get("Padj"))
        or to_float(row.get("FDR"))
    )
    n = to_int(row.get("N"))
    effect_percent = round((1 - hr) * 100, 3) if hr is not None else None
    is_higher_acm = effect_percent is not None and effect_percent < 0
    direction = "higher" if is_higher_acm else "lower"
    ci = f"{fmt_float(lower)}-{fmt_float(upper)}"
    table = str(row["source_table"])
    source_kind = str(row["source_kind"])
    is_fdr_significant = is_true(row.get("is_fdr_significant"))
    is_statistically_significant = is_true(row.get("is_statistically_significant"))
    dataset_link = SUPPLEMENT_DATASET_LINKS.get(table, SUPPLEMENT_DATASET_LINKS["Data Table 2"])
    group = f"HUMAN_ACM_{slugify(display)}"
    result_detail = (
        f"Public Aging Cell supplement {table} ({source_label(table)}): "
        f"N={n:,} prescription users; HR {fmt_float(hr)} "
        f"(95% CI {ci}); P={format_p(p_value)}"
        f"{'; adjusted/FDR value ' + format_p(fdr_value) if fdr_value is not None else ''}."
    )
    row_note = ROW_NOTES.get(display)
    default_row_note = (
        "This Figure 5 / Data Table 4 row is shown alongside the full 406-medication N>=500 screen."
        if source_kind == "class"
        else "This row is shown as part of the full 406-medication N>=500 screen."
    )
    followup_detail = (
        f"The row is retained from Figure 5 / Data Table 4 and reports {direction} "
        f"all-cause mortality. Source type: {source_kind}."
        if source_kind == "class"
        else (
            f"The row is retained from the full 406-medication Data Table 2 N>=500 "
            f"screen and reports {direction} all-cause mortality. "
            f"Source type: {source_kind}."
        )
    )

    meta = {
        "cohort": COHORT,
        "group": group,
        "control_group": None,
        "is_control": False,
        "label": display,
        "description": display_info["description"],
        "species": "Human",
        "strain": display_info["condition"],
        "compound_display_name": display,
        "condition": display_info["condition"],
        "publication": PUBLICATION,
        "evidence": {
            "overall": {
                "score": effect_percent,
                "value_kind": "acm_effect",
                "label": overall_effect_label(
                    sort_index,
                    source_kind,
                    effect_percent,
                    is_fdr_significant or is_statistically_significant,
                ),
                "note": row_note or default_row_note,
            },
            "biobank": {
                "score": 100,
                "effect_label": biobank_effect_label(display, source_kind, hr),
                "detail": result_detail,
                "publication": PUBLICATION,
            },
            "followup": {
                "score": min(100, max(35, round(math.log10(max(n or 1, 1)) * 22))),
                "effect_label": (
                    "Cohort-support score derived from the number of UK Biobank prescription users in the public supplement row."
                ),
                "detail": followup_detail,
                "publication": PUBLICATION,
            },
            "caution": {
                "value": "Observational",
                "value_kind": "text",
                "label": "Retrospective association",
                "effect_label": (
                    "Retrospective UK Biobank association, not a randomized treatment-effect estimate."
                ),
                "detail": (
                    "Interpret with indication, prescribing, sex, disease-severity, and healthy-user confounding in mind. "
                    "The site ships only public supplement and metadata summaries, not participant-level UK Biobank data."
                ),
                "publication": PUBLICATION,
            },
        },
        "acm_hazard_ratio": round_float(hr, 6),
        "acm_confidence_interval": ci,
        "acm_effect_percent": effect_percent,
        "acm_p_value": p_value,
        "acm_adjusted_p_value": fdr_value,
        "acm_prescription_user_count": n,
        "acm_source_kind": source_kind,
        "acm_source_table": table,
        "acm_fdr_significant": is_fdr_significant,
        "acm_statistically_significant": is_statistically_significant,
        "acm_dataset_label": dataset_link["label"],
        "acm_dataset_url": dataset_link["url"],
        "acm_dataset_detail": dataset_link["detail"],
        "source_catalog": SOURCE_CATALOG,
        "raw_result": {
            "source_table": table,
            "source_kind": source_kind,
            "dataset_label": dataset_link["label"],
            "dataset_url": dataset_link["url"],
            "raw_drug": raw_drug,
            "n": n,
            "hr": round_float(hr, 9),
            "lower95": round_float(lower, 9),
            "upper95": round_float(upper, 9),
            "p": p_value,
            "fdr_or_padj": fdr_value,
            "padj_n500": to_float(row.get("Padj N>500")),
            "rank_n500": to_int(row.get("Rank N>500")),
            "fdr_significant": is_fdr_significant,
            "statistically_significant": is_statistically_significant,
        },
        "sort_index": sort_index,
    }
    return group, meta


def build_manifest() -> dict[str, Any]:
    if not WORKBOOK_PATH.exists():
        raise SystemExit(
            f"Missing {WORKBOOK_PATH.relative_to(ROOT)}. Run npm run download:human:acm first.",
        )

    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    rows = screened_acm_rows(workbook)
    rows.sort(
        key=lambda row: (
            to_float(row.get("HR")) or 99,
            str(row["Drug"]),
        ),
    )

    group_entries = [build_group_meta(row, sort_index=index + 1) for index, row in enumerate(rows)]
    group_meta_by_key = {f"{COHORT}::{group}": meta for group, meta in group_entries}
    first_group = group_entries[0][0] if group_entries else None
    raw_counts = {
        "screened_no_concentration_drug_rows": sum(
            1
            for row in rows_from_sheet(workbook, "Data Table 2")
            if (to_int(row.get("N")) or 0) >= 500
        ),
        "rendered_acm_rows": len(rows),
        "lower_acm_rows": sum(1 for row in rows if (to_float(row.get("HR")) or 0) < 1),
        "higher_acm_rows": sum(1 for row in rows if (to_float(row.get("HR")) or 0) > 1),
        "class_acm_rows": sum(1 for row in rows if row["source_table"] == "Data Table 4"),
        "data_table_2_rows": sum(1 for row in rows if row["source_table"] == "Data Table 2"),
        "data_table_3_rows": sum(1 for row in rows if row["source_table"] == "Data Table 3"),
        "data_table_4_rows": sum(1 for row in rows if row["source_table"] == "Data Table 4"),
    }

    return {
        "generated_at_utc": utc_now_iso(),
        "source_scope": (
            "Public Aging Cell supplement source pool: all 406 Data Table 2 no-concentration "
            "medication ACM rows with N>=500, plus the six Figure 5 / Data Table 4 "
            "medication-class rows. The site view displays rows with reported P values "
            "and ranks them as 1 minus the reported ACM hazard ratio. Includes public UK "
            "Biobank project provenance; no participant-level UK Biobank data."
        ),
        "raw_workbook": str(WORKBOOK_PATH.relative_to(ROOT)),
        "download_manifest": str(DOWNLOAD_MANIFEST_PATH.relative_to(ROOT))
        if DOWNLOAD_MANIFEST_PATH.exists()
        else None,
        "latest_public_cohort_downloaded": COHORT,
        "latest_public_release_label": "Aging Cell 2024 supplement",
        "cohort_order": [COHORT],
        "site_meta": {
            "ALL": "UK Biobank public supplement",
        },
        "profile": {
            "title": "Human UK Biobank ACM Interventions",
            "description": (
                "A Biobank-only human view of ACM rows with reported P values from all 406 "
                "Data Table 2 N>=500 medication rows plus Figure 5 medication-class rows, "
                "ranked by the reported hazard-ratio effect."
            ),
            "interventionStatLabel": "Signals",
            "cohortStatLabel": "Study",
            "groupNounSingular": "ACM signal",
            "groupNounPlural": "ACM signals",
            "focusGroupLabel": "Focus signal",
            "compareSectionLabel": "Compare Biobank signals",
            "explorerSectionLabel": "Biobank signal explorer",
            "defaultRankingMetric": "acm",
            "allowCompare": False,
            "defaultGroup": first_group,
        },
        "default_group": first_group,
        "source_catalog": SOURCE_CATALOG,
        "cohort_meta_by_name": {
            COHORT: {
                "label": "UK Biobank ACM screen (Morin et al. 2024)",
                "short_label": "UKB ACM 2024",
                "secondary_label": "Aging Cell supplement",
                "publication": PUBLICATION,
            },
        },
        "raw_counts": raw_counts,
        "group_meta_by_key": group_meta_by_key,
    }


def main() -> None:
    META_DIR.mkdir(parents=True, exist_ok=True)
    manifest = build_manifest()
    OUTPUT_PATH.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    counts = manifest["raw_counts"]
    print(
        "Wrote "
        f"{OUTPUT_PATH.relative_to(ROOT)} with {counts['rendered_acm_rows']} ACM rows "
        f"from {counts['screened_no_concentration_drug_rows']} screened no-concentration drugs "
        f"plus {counts['class_acm_rows']} class rows "
        f"({counts['lower_acm_rows']} lower, {counts['higher_acm_rows']} higher).",
    )


if __name__ == "__main__":
    main()
